import PyPDF2
import pdfplumber
import re
import pandas as pd
import os
import glob
from collections import defaultdict
from typing import List, Dict

class BulkPOItemExtractor:
    def __init__(self, folder_path: str):
        self.folder_path = folder_path
        self.all_pos_data = {}
        self.combined_items = defaultdict(lambda: {'total_quantity': 0, 'po_files': [], 'quantities_per_po': {}})

    def get_pdf_files(self) -> List[str]:
        pdf_files = glob.glob(os.path.join(self.folder_path, "*.pdf"))
        print(f"Found {len(pdf_files)} PDF files" if pdf_files else f"No PDF files found in {self.folder_path}")
        return pdf_files

    def extract_text_pdfplumber(self, pdf_path: str) -> str:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                full_text = ""
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    tables = page.extract_tables()
                    if tables:
                        for table in tables:
                            for row in table[1:]:  # Skip header
                                if row and len(row) > 2 and re.match(r'^\d{5}$', str(row[0])):
                                    text += "\n".join(str(cell) for cell in row if cell) + "\n"
                    full_text += text
                return full_text if full_text else "\n".join(page.extract_text() or "" for page in pdf.pages)
        except Exception as e:
            print(f"Error extracting from {os.path.basename(pdf_path)} with pdfplumber: {e}")
            return ""

    def extract_text_pypdf2(self, pdf_path: str) -> str:
        try:
            with open(pdf_path, 'rb') as file:
                return "".join(page.extract_text() or "" for page in PyPDF2.PdfReader(file).pages)
        except Exception as e:
            print(f"Error extracting from {os.path.basename(pdf_path)} with PyPDF2: {e}")
            return ""

    def extract_po_metadata(self, text: str) -> Dict[str, str]:
        metadata = {}
        patterns = {
            'po_date': r'PO Date:\s*(\d{2}\.\d{2}\.\d{4})',
            'document_ref': r'Document Ref:\s*(\d+)',
            'vendor': r'Vendor.*?([A-Za-z\s]+)(?=\s*Code:|\n)',
            'total_amount': r'Total Including Sales Tax\s*([\d,]+\.?\d*)'
        }
        for key, pattern in patterns.items():
            match = re.search(pattern, text, re.DOTALL if key == 'vendor' else 0)
            if match:
                metadata[key] = match.group(1).strip() if key == 'vendor' else match.group(1)
        return metadata

    def parse_items_from_text(self, text: str) -> List[Dict]:
        items = []
        lines = [line.strip() for line in text.split('\n') if line.strip()]
        print(f"Parsing text with {len(lines)} non-empty lines...")
        i = 0
        while i < len(lines):
            line = lines[i]
            print(f"Line {i}: {line}")
            item_match = re.match(r'^(\d{5})\s+(.+?)\s+(\d+\.?\d*)\s+\d+\.?\d*\s+\d+\.?\d*\s+\d+\.?\d*\s+\d+\.?\d*$', line)
            if item_match:
                item_number = item_match.group(1)
                name = item_match.group(2).strip()
                qty = float(item_match.group(3))
                print(f"  Found item: {item_number}, {name}, {qty}")
                items.append({'item_number': item_number, 'name': name, 'quantity': qty})
                i += 1  # Move to next line after finding an item
                continue
            i += 1
        print(f"Extracted {len(items)} items.")
        return items

    def get_short_po_name(self, po_file: str) -> str:
        match = re.search(r'LHR\s*([A-Z0-9\s]+?)\s*Stock\.pdf$', po_file, re.IGNORECASE)
        if match:
            return f"LHR{match.group(1).strip().replace(' ', '')}"
        return po_file.replace('.pdf', '')[:20].replace(' ', '_')

    def process_single_pdf(self, pdf_path: str) -> Dict:
        filename = os.path.basename(pdf_path)
        print(f"Processing: {filename}")
        
        text = self.extract_text_pdfplumber(pdf_path)
        if not text:
            text = self.extract_text_pypdf2(pdf_path)
        if not text:
            print(f"  ❌ Failed to extract text from {filename}")
            return {'filename': filename, 'success': False, 'items': [], 'metadata': {}}
        
        metadata = self.extract_po_metadata(text)
        items = self.parse_items_from_text(text)
        print(f"  ✅ Extracted {len(items)} items from {filename}" if items else f"  ⚠️ No items found in {filename}")
        
        return {'filename': filename, 'success': bool(items), 'items': items, 'metadata': metadata}

    def process_all_pdfs(self):
        for pdf_path in self.get_pdf_files():
            result = self.process_single_pdf(pdf_path)
            self.all_pos_data[result['filename']] = result
            if result['success'] and result['items']:
                for item in result['items']:
                    item_key = item['name'].strip()
                    self.combined_items[item_key]['total_quantity'] += item['quantity']
                    self.combined_items[item_key]['po_files'].append(result['filename'])
                    self.combined_items[item_key]['quantities_per_po'][result['filename']] = item['quantity']
            print()

    def print_individual_po_summary(self):
        print(f"\n{'='*80}\nINDIVIDUAL PO SUMMARIES\n{'='*80}")
        for filename, data in self.all_pos_data.items():
            print(f"\n📄 File: {filename}")
            for key, emoji in [('po_date', '📅'), ('document_ref', '📋'), ('vendor', '🏢'), ('total_amount', '💰')]:
                if key in data['metadata']:
                    print(f"   {emoji} {key.replace('_', ' ').title()}: {'PKR ' if key == 'total_amount' else ''}{data['metadata'][key]}")
            if data['success'] and data['items']:
                print(f"   📦 Items: {len(data['items'])}\n   {'Item #':<8} {'Qty':<10} {'Item Name'}\n   {'-'*70}")
                for item in data['items']:
                    print(f"   {item['item_number']:<8} {item['quantity']:<10.0f} {item['name'][:50]}")
            else:
                print(f"   ❌ No items extracted")

    def create_procurement_table(self):
        print(f"\n{'='*150}\n📋 DETAILED PROCUREMENT TABLE\n{'='*150}")
        if not self.combined_items:
            print("No items found across all POs")
            return
        all_po_files = sorted(set().union(*(data['po_files'] for data in self.combined_items.values())))
        table_data = [
            {'Item_Name': name, 'Total_Quantity': data['total_quantity'], **{self.get_short_po_name(po): data['quantities_per_po'].get(po, 0) for po in all_po_files}}
            for name, data in sorted(self.combined_items.items())
        ]
        df = pd.DataFrame(table_data).fillna(0)
        print(f"\n📊 Items: {len(table_data)} | POs: {len(all_po_files)}\n\n{'='*150}")
        po_headers = [self.get_short_po_name(po) for po in all_po_files]
        header = f"{'#':<3} {'Item Name':<50} {'TOTAL':<10} {' '.join(f'{h:<15}' for h in po_headers)}"
        print(header + "\n" + "-" * len(header))
        for idx, row in df.iterrows():
            item_name = row['Item_Name'][:48] + ".." if len(row['Item_Name']) > 50 else row['Item_Name']
            line = f"{idx+1:<3} {item_name:<50} {row['Total_Quantity']:<10.0f}"
            line += "".join(f" {row[po]:<15.0f}" if row[po] > 0 else f" {'-':<15}" for po in po_headers)
            print(line)
        print("-" * len(header))
        print(f"{'TOT':<3} {'All Items Total':<50} {df['Total_Quantity'].sum():<10.0f}" + 
              "".join(f" {df[po].sum():<15.0f}" for po in po_headers))
        print("=" * len(header))
        return df

    def save_procurement_table(self, output_file: str = 'procurement_table.xlsx'):
        try:
            all_po_files = sorted(set().union(*(data['po_files'] for data in self.combined_items.values())))
            table_data = [
                {'Item_Name': name, 'Total_Quantity': data['total_quantity'], **{f'PO_{self.get_short_po_name(po)}': data['quantities_per_po'].get(po, 0) for po in all_po_files}}
                for name, data in sorted(self.combined_items.items())
            ]
            df = pd.DataFrame(table_data).fillna(0)
            output_path = os.path.join(self.folder_path, output_file)
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Procurement Table', index=False)
                worksheet = writer.sheets['Procurement Table']
                from openpyxl.styles import Font, PatternFill, Alignment
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                worksheet.column_dimensions['A'].width = 60
                worksheet.column_dimensions['B'].width = 15
                for col_idx in range(3, len(df.columns) + 1):
                    worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 15
            print(f"\n📊 Procurement table saved to: {output_path}")
        except Exception as e:
            print(f"Error saving procurement table: {e}")

    def save_to_excel(self, output_file: str = 'po_analysis.xlsx'):
        try:
            with pd.ExcelWriter(os.path.join(self.folder_path, output_file), engine='openpyxl') as writer:
                pd.DataFrame([
                    {'Item Name': name, 'Total Quantity': data['total_quantity'], 'Number of POs': len(data['po_files']), 'PO Files': ', '.join(self.get_short_po_name(f) for f in data['po_files'])}
                    for name, data in self.combined_items.items()
                ]).sort_values('Total Quantity', ascending=False).to_excel(writer, sheet_name='Combined Summary', index=False)
                pd.DataFrame([
                    {'PO File': self.get_short_po_name(f), 'Success': d['success'], 'Number of Items': len(d['items']) if d['success'] else 0, **{k: d['metadata'].get(k, '') for k in ['po_date', 'document_ref', 'vendor', 'total_amount']}}
                    for f, d in self.all_pos_data.items()
                ]).to_excel(writer, sheet_name='PO Summary', index=False)
                pd.DataFrame([
                    {'PO File': self.get_short_po_name(f), 'Item Number': i['item_number'], 'Item Name': i['name'], 'Quantity': i['quantity'], 'PO Date': d['metadata'].get('po_date', ''), 'Vendor': d['metadata'].get('vendor', '')}
                    for f, d in self.all_pos_data.items() if d['success'] and d['items'] for i in d['items']
                ]).to_excel(writer, sheet_name='All Items Detail', index=False)
                # New Quantity Summary Sheet
                summary_data = []
                for item_name, data in self.combined_items.items():
                    row = {'Row Labels': item_name}
                    for po_file in self.all_pos_data.keys():
                        po_short_name = self.get_short_po_name(po_file)
                        row[po_short_name] = data['quantities_per_po'].get(po_file, 0)
                    row['Grand Total'] = data['total_quantity']
                    summary_data.append(row)
                summary_df = pd.DataFrame(summary_data).fillna(0)
                summary_df = summary_df[['Row Labels'] + [self.get_short_po_name(f) for f in self.all_pos_data.keys()] + ['Grand Total']]
                summary_df.to_excel(writer, sheet_name='Quantity Summary', index=False)
                summary_worksheet = writer.sheets['Quantity Summary']
                for cell in summary_worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                summary_worksheet.column_dimensions['A'].width = 50
                for col_idx in range(2, len(summary_df.columns) + 1):
                    summary_worksheet.column_dimensions[summary_worksheet.cell(row=1, column=col_idx).column_letter].width = 15
            print(f"\n💾 Data saved to: {os.path.join(self.folder_path, output_file)}")
        except Exception as e:
            print(f"Error saving to Excel: {e}")

    def run_analysis(self):
        print(f"🚀 Starting Bulk PO Analysis...\n📁 Folder: {self.folder_path}")
        self.process_all_pdfs()
        self.print_individual_po_summary()
        self.create_procurement_table()
        self.save_to_excel()
        self.save_procurement_table()
        total_pos = sum(1 for data in self.all_pos_data.values() if data['success'])
        total_unique_items = len(self.combined_items)
        total_quantity = sum(data['total_quantity'] for data in self.combined_items.values())
        print(f"\n{'='*60}\n📈 FINAL STATISTICS\n{'='*60}")
        print(f"✅ Successfully processed POs: {total_pos}")
        print(f"📦 Total unique items: {total_unique_items}")
        print(f"🔢 Total quantity across all POs: {total_quantity:.0f}")
        print(f"💾 Results saved to Excel file")

def main():
    folder_path = r"C:\Users\Hassan Shahzad\Downloads\PO"
    if not os.path.exists(folder_path):
        print(f"❌ Folder not found: {folder_path}\nPlease make sure the folder path is correct.")
        return
    BulkPOItemExtractor(folder_path).run_analysis()

if __name__ == "__main__":
    main()